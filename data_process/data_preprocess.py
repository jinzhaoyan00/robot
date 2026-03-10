"""
Extract text data from Excel files in data/知识库-v2/ and save as JSON to data/txt/.

Processes three sheets per file:
  - 配置:    header row + data rows → 2-row markdown tables; last row → text block
  - 答疑:    header row + data rows → 2-row markdown tables
  - 电池保修: first N rows → full markdown table; last 2 rows → text blocks
"""

import json
import os
import uuid
from pathlib import Path

import openpyxl
from openpyxl.utils import range_boundaries


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_merge_map(ws) -> dict:
    """Return {(row, col): value} for every cell that belongs to a merged range."""

    class BreakToOuter(Exception):
        pass

    break_flag = False
    merge_map: dict = {}
    for merged_range in ws.merged_cells.ranges:
        try:
            value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merge_map[(r, c)] = value
                    if isinstance(value, str) and len(value) > 160:   # 满足条件时跳出到最外层
                        raise BreakToOuter
        except BreakToOuter:
            continue 

    return merge_map


def _cell_value(ws, row: int, col: int, merge_map: dict):
    return merge_map.get((row, col), ws.cell(row=row, column=col).value)


def _fmt(v) -> str:
    if v is None:
        return ""
    return str(v).replace('\n', '\\n').strip()


def _read_rows(ws, merge_map: dict) -> list[list]:
    """Read all rows, expanding merged cells."""
    rows = []
    for r in range(1, ws.max_row + 1):
        rows.append([_cell_value(ws, r, c, merge_map) for c in range(1, ws.max_column + 1)])
    return rows


def _make_2row_table(headers: list, data_row: list) -> str:
    """Build a markdown table with exactly 2 data rows (header + one data row)."""
    n = len(headers)
    h = [_fmt(v) for v in headers]
    d = [_fmt(v) for v in data_row[:n]] + [""] * max(0, n - len(data_row))
    sep = ["---"] * n
    lines = [
        "| " + " | ".join(h) + " |",
        "| " + " | ".join(sep) + " |",
        "| " + " | ".join(d) + " |",
    ]
    return "\n".join(lines)


def _make_full_table(rows: list[list]) -> str:
    """Build a markdown table from multiple rows (first row = header)."""
    if not rows:
        return ""
    n = max(len(r) for r in rows)
    padded = [[_fmt(v) for v in r] + [""] * (n - len(r)) for r in rows]
    sep = ["---"] * n
    lines = ["| " + " | ".join(padded[0]) + " |",
             "| " + " | ".join(sep) + " |"]
    lines += ["| " + " | ".join(row) + " |" for row in padded[1:]]
    return "\n".join(lines)


def _row_to_text(row: list) -> str:
    """Concatenate non-empty cell values from a row into a single string."""
    return "\n".join(part for v in row if (part := _fmt(v)))


def _make_record(document: str, category: str, sheet_name: str, idx: int) -> dict:
    return {
        "document": document.replace(" |  |  |  |  |", ""),
        "metadata": {
            "category": category,
            "tags": [sheet_name, category],
        },
        "id": f"{category}_{sheet_name}_{idx}",
    }


# ---------------------------------------------------------------------------
# Sheet processors
# ---------------------------------------------------------------------------

def process_peizhi(ws, category: str) -> list[dict]:
    """
    配置 sheet:
      - Row 1          → header
      - Rows 2 .. N-1  → data rows  →  2-row markdown table each
      - Row N          → notes text block
    """
    merge_map = _build_merge_map(ws)
    rows = _read_rows(ws, merge_map)
    if len(rows) < 2:
        return []

    header = rows[0]
    data_rows = rows[1:-1]
    notes_row = rows[-1]

    records: list[dict] = []
    for i, data_row in enumerate(data_rows):
        doc = _make_2row_table(header, data_row)
        records.append(_make_record(doc, category, "配置", i))

    notes_text = _row_to_text(notes_row)
    if notes_text:
        records.append(_make_record(notes_text, category, "配置", len(data_rows)))

    return records


def process_dayi(ws, category: str) -> list[dict]:
    """
    答疑 sheet:
      - Row 1        → header
      - Rows 2 .. N  → data rows  →  2-row markdown table each
    """
    merge_map = _build_merge_map(ws)
    rows = _read_rows(ws, merge_map)
    if len(rows) < 2:
        return []

    header = rows[0]
    data_rows = rows[1:]

    records: list[dict] = []
    for i, data_row in enumerate(data_rows):
        doc = _make_2row_table(header, data_row)
        records.append(_make_record(doc, category, "答疑", i))

    return records


def process_battery(ws, category: str) -> list[dict]:
    """
    电池保修 sheet:
      - Rows 1 .. N-2  → one full markdown table
      - Row N-1        → text block
      - Row N          → text block
    """
    rows = [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]
    if len(rows) < 3:
        return []

    table_rows = rows[:-2]
    last_two = rows[-2:]

    records: list[dict] = []
    table_text = _make_full_table(table_rows)
    if table_text:
        records.append(_make_record(table_text, category, "电池保修", 0))

    for i, row in enumerate(last_two, start=1):
        text = _row_to_text(row)
        if text:
            records.append(_make_record(text, category, "电池保修", i))

    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

SHEET_PROCESSORS = {
    "配置": process_peizhi,
    "答疑": process_dayi,
    "电池保修": process_battery,
}


def trim_workbook(wb):
    """
    删除工作簿中每个工作表最下面和最右面的空白行和列。
    :param wb: 已打开的 openpyxl.Workbook 对象
    """
    for ws in wb.worksheets:
        # 计算实际数据区域的维度（考虑有值的单元格和合并单元格）
        dim = ws.calculate_dimension()
        min_col, min_row, max_col, max_row = range_boundaries(dim)

        # 删除右侧多余的列（从数据区域最后一列+1 到工作表最大列）
        if max_col < ws.max_column:
            ws.delete_cols(max_col + 1, ws.max_column - max_col)

        # 删除下方多余的行（从数据区域最后一行+1 到工作表最大行）
        if max_row < ws.max_row:
            ws.delete_rows(max_row + 1, ws.max_row - max_row)


def process_excel(excel_path: Path, output_dir: Path) -> None:
    category = excel_path.stem
    wb = openpyxl.load_workbook(excel_path)
    trim_workbook(wb)

    for sheet_name, processor in SHEET_PROCESSORS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] sheet '{sheet_name}' not found in {excel_path.name}")
            continue

        ws = wb[sheet_name]
        records = processor(ws, category)

        out_path = output_dir / f"{category}_{sheet_name}.json"
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)

        print(f"  Wrote {len(records):>4} records → {out_path.name}")


# {'min': 60, 'max': 761, 'average': 201.77553083923155}
def calculate_document_length_stats(folder_path = '/data/txt'):
    """
    计算指定文件夹下所有JSON文件中每个"document"内容的长度，
    返回包含最小值、最大值和平均值的字典。
    """
    
    lengths = []

    # 遍历文件夹
    for filename in os.listdir(folder_path):
        if not filename.endswith('.json'):
            continue

        file_path = os.path.join(folder_path, filename)
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

                # 处理列表结构（如示例文件）
                if isinstance(data, list):
                    for item in data:
                        if isinstance(item, dict) and 'document' in item:
                            lengths.append(len(item['document']))
                # 处理单对象结构（如有）
                elif isinstance(data, dict) and 'document' in data:
                    lengths.append(len(data['document']))

        except (json.JSONDecodeError, IOError) as e:
            print(f"处理文件 {filename} 时出错: {e}")
            continue

    # 计算统计值
    if lengths:
        min_len = min(lengths)
        max_len = max(lengths)
        avg_len = sum(lengths) / len(lengths)
        return {'min': min_len, 'max': max_len, 'average': avg_len}
    else:
        return {'min': None, 'max': None, 'average': None}


def main() -> None:
    base_dir = Path(__file__).resolve().parent.parent
    input_dir = base_dir / "data" / "知识库-v2"
    output_dir = base_dir / "data" / "txt"
    output_dir.mkdir(parents=True, exist_ok=True)

    excel_files = sorted(input_dir.glob("*.xlsx"))
    if not excel_files:
        print(f"No .xlsx files found in {input_dir}")
        return

    for excel_path in excel_files:
        print(f"\nProcessing: {excel_path.name}")
        process_excel(excel_path, output_dir)

    res = calculate_document_length_stats(output_dir)
    print(res)
    print("\nDone.")


if __name__ == "__main__":
    main()
